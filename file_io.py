import openpyxl

def read_word_vector(file_path):
    '''
    param
        file_path: file of word vector file
    return:
        words: list of all words
        vectors: list of all word vectors
    '''

    file = open(file_path)
    line = file.readline().split(' ')
    num_words = int(line[0])
    size_vector = int(line[1])
    print('Number of words: %d, size of vector: %d' %(num_words, size_vector))
    vectors = []
    words = []
    for i in range(num_words):
        if i % int(num_words/10) == 0:
            print('%d%%' % (i*100/num_words))
        line = file.readline().split(' ')
        word = line[0]
        vector = [float(i) for i in line[1:]]
        words.append(word)
        vectors.append(vector)
    file.close()
    return words, vectors


def write_word_vector(words, labels, n_clusters, n_words):
    if len(words) != len(labels):
        print('Numbers of words and labels don\'t match!')
    else:
        file = open('./cluster-%d-%d-words.txt' % (n_clusters, n_words), 'w')
        for i in range(0, len(words)):
            file.write('%d\t%s\n' % (labels[i], words[i]))
        file.close()


def read_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    a_sheet = wb.get_sheet_by_name('Sheet1')
    pair_list = []
    for i in range(1, 51):
        cell = a_sheet.cell(row=i, column=1).value.split('\n\n')
        pair = {'questions': cell[1].split('\n'),
                'answer': cell[2]}
        pair_list.append(pair)
    return pair_list


def write_excel(pair_list, save_path):
    num = len(pair_list)
    wb = openpyxl.Workbook()
    wb.create_sheet('Sheet1', index=0)
    sheet = wb.get_sheet_by_name('Sheet1')
    sheet.title = 'Sheet1'
    row = ['ID', 'Answer', 'Category', 'Keywords', 'From date', 'To date', 'Question', 'Question', 'More Questions']
    sheet.append(row)
    for i in range(2, num+2):
        pair = pair_list[i-2]
        row = [i-1, pair['answer'], '', '', '', '', pair['questions'][0], '']
        if len(pair['questions']) > 1:
            row[7] = pair['questions'][1]
        sheet.append(row)
    wb.save(save_path)


if __name__ == '__main__':
    file_path = './data/faq_origin.xlsx'
    save_path = './result/faq_pairs.xlsx'
    # read_word_vector(file_path)
    pair_list = read_excel(file_path)
    write_excel(pair_list, save_path)